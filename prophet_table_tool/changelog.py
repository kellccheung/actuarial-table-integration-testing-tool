"""Stage 1 – Generate Change Log from before/ vs after/ tables."""

from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .audit import close_audit_logger, setup_audit_logger
from .changelog_review import TableReviewModel, add_contribution, write_review_sheets
from .control import read_control
from .diff import ChangeRow, detect_conflicts, diff_table
from .prophet_csv import discover_csv_tables, read_prophet_csv
from .utils import file_hash, generate_run_id


DETAIL_HEADERS = [
    "change_request_id",
    "table_name",
    "change_type",
    "n_keys_before",
    "n_keys_after",
    "key_tuple",
    "column_name",
    "old_value",
    "new_value",
    "notes",
]

SUMMARY_HEADERS = [
    "change_request_id",
    "description",
    "tables_touched",
    "n_value_changes",
    "n_row_adds",
    "n_row_deletes",
    "n_column_changes",
    "n_key_count_changes",
    "has_conflict",
    "status",
]

CONFLICT_HEADERS = [
    "conflict_type",
    "table_name",
    "key_tuple",
    "column_name",
    "change_request_ids",
    "change_types",
    "old_values",
    "new_values",
    "notes",
    "resolved",
]


def generate_change_log(control_path: Path) -> Path:
    """
    Stage 1: Compare before/ vs after/ for each included change request
    and write a ChangeLog_*.xlsx plus ChangeLog_*_Detail.csv sidecar.

    Returns the path to the generated Change Log workbook.
    """
    control_path = Path(control_path)
    control = read_control(control_path)
    run_id = control.run_id or generate_run_id()
    control.output_path.mkdir(parents=True, exist_ok=True)

    logger, _log_path = setup_audit_logger(
        control.output_path, run_id, "generate_changelog"
    )
    status = "FAILED"
    out_path: Path | None = None

    try:
        logger.info("control_path=%s", control_path)
        logger.info("control_hash=%s", file_hash(control_path))
        logger.info("working_root=%s", control.working_root)

        included = control.included_requests(require_approved=False)
        logger.info(
            "change_requests_processed=%s",
            [r.change_request_id for r in included],
        )

        all_changes: list[ChangeRow] = []
        summary_rows: list[dict] = []
        all_warnings: list[str] = []
        table_reviews: dict[str, TableReviewModel] = {}

        for cr in included:
            cr_dir = control.change_request_dir(cr.change_request_id)
            before_dir = cr_dir / "before"
            after_dir = cr_dir / "after"

            logger.info("[%s] discovering tables in before/ and after/", cr.change_request_id)
            before_tables = discover_csv_tables(before_dir)
            after_tables = discover_csv_tables(after_dir)
            logger.info(
                "[%s] found %d before table(s), %d after table(s)",
                cr.change_request_id,
                len(before_tables),
                len(after_tables),
            )

            if not before_tables and not after_tables:
                msg = (
                    f"[{cr.change_request_id}] Empty change request folder "
                    f"(no CSVs in before/ or after/)."
                )
                logger.warning(msg)
                all_warnings.append(msg)
                summary_rows.append(
                    {
                        "change_request_id": cr.change_request_id,
                        "description": cr.description,
                        "tables_touched": "",
                        "n_value_changes": 0,
                        "n_row_adds": 0,
                        "n_row_deletes": 0,
                        "n_column_changes": 0,
                        "n_key_count_changes": 0,
                        "has_conflict": "N",
                        "status": "EMPTY",
                    }
                )
                continue

            only_before = set(before_tables) - set(after_tables)
            for table_name in sorted(only_before):
                msg = (
                    f"[{cr.change_request_id}] Table '{table_name}' exists only "
                    "in before/ — skipped (no automatic deletes)."
                )
                logger.warning(msg)
                all_warnings.append(msg)

            cr_changes: list[ChangeRow] = []
            tables_touched: set[str] = set()

            all_names = set(before_tables) | set(after_tables)
            for table_name in sorted(all_names):
                if table_name in only_before:
                    continue

                before_tbl = (
                    read_prophet_csv(before_tables[table_name])
                    if table_name in before_tables
                    else None
                )
                after_tbl = (
                    read_prophet_csv(after_tables[table_name])
                    if table_name in after_tables
                    else None
                )

                n_before = 0 if before_tbl is None else before_tbl.data.height
                n_after = 0 if after_tbl is None else after_tbl.data.height
                if before_tbl is not None and before_tbl.source_encoding:
                    logger.info(
                        "[%s] read before %s encoding=%s",
                        cr.change_request_id,
                        table_name,
                        before_tbl.source_encoding,
                    )
                if after_tbl is not None and after_tbl.source_encoding:
                    logger.info(
                        "[%s] read after %s encoding=%s",
                        cr.change_request_id,
                        table_name,
                        after_tbl.source_encoding,
                    )
                logger.info(
                    "[%s] comparing %s (%s→%s rows)...",
                    cr.change_request_id,
                    table_name,
                    n_before,
                    n_after,
                )

                diff = diff_table(
                    cr.change_request_id,
                    table_name,
                    before_tbl,
                    after_tbl,
                    control,
                )
                cr_changes.extend(diff.change_rows)
                tables_touched |= diff.tables_touched
                logger.info(
                    "[%s] %s done, %d change row(s)",
                    cr.change_request_id,
                    table_name,
                    len(diff.change_rows),
                )
                if diff.change_rows or table_name in diff.tables_touched:
                    add_contribution(
                        table_reviews,
                        table_name,
                        cr.change_request_id,
                        before_tbl,
                        after_tbl,
                        diff.change_rows,
                    )
                for w in diff.warnings:
                    logger.warning(w)
                    all_warnings.append(w)

            all_changes.extend(cr_changes)
            counts = Counter(r.change_type for r in cr_changes)
            n_col = (
                counts.get("column_add", 0)
                + counts.get("column_delete", 0)
                + counts.get("column_rename", 0)
            )
            summary_rows.append(
                {
                    "change_request_id": cr.change_request_id,
                    "description": cr.description,
                    "tables_touched": ", ".join(sorted(tables_touched)),
                    "n_value_changes": counts.get("value_update", 0),
                    "n_row_adds": counts.get("row_add", 0),
                    "n_row_deletes": counts.get("row_delete", 0),
                    "n_column_changes": n_col,
                    "n_key_count_changes": counts.get("key_count_change", 0),
                    "has_conflict": "N",
                    "status": "OK" if cr_changes or tables_touched else "NO_CHANGES",
                }
            )

        conflicts = detect_conflicts(all_changes)
        conflicted_crs: set[str] = set()
        for c in conflicts:
            conflicted_crs.update(c["change_request_ids"])

        for row in summary_rows:
            if row["change_request_id"] in conflicted_crs:
                row["has_conflict"] = "Y"
                if row["status"] == "OK":
                    row["status"] = "CONFLICT"

        out_path = control.output_path / f"ChangeLog_{run_id}.xlsx"
        detail_csv_path = control.output_path / f"ChangeLog_{run_id}_Detail.csv"

        logger.info("writing detail CSV (%d rows): %s", len(all_changes), detail_csv_path)
        _write_change_log_detail_csv(detail_csv_path, all_changes)

        logger.info("writing Change Log workbook (Summary/Conflicts/reviews): %s", out_path)
        _write_change_log_excel(
            out_path, summary_rows, conflicts, table_reviews
        )

        logger.info("tables_affected=%d", len({r.table_name for r in all_changes}))
        logger.info("n_change_rows=%d", len(all_changes))
        logger.info("n_conflicts=%d", len(conflicts))
        for w in all_warnings:
            logger.info("warning=%s", w)
        logger.info("change_log_path=%s", out_path)
        logger.info("change_log_hash=%s", file_hash(out_path))
        logger.info("change_log_detail_path=%s", detail_csv_path)
        logger.info("change_log_detail_hash=%s", file_hash(detail_csv_path))

        status = "SUCCESS"
        logger.info("final_status=%s", status)
        return out_path

    except Exception as exc:
        logger.exception("generate_change_log failed: %s", exc)
        logger.info("final_status=%s", status)
        raise
    finally:
        close_audit_logger(logger)


def _write_change_log_detail_csv(path: Path, change_rows: list[ChangeRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(DETAIL_HEADERS)
        for row in change_rows:
            writer.writerow(
                [
                    row.change_request_id,
                    row.table_name,
                    row.change_type,
                    row.n_keys_before if row.n_keys_before is not None else "",
                    row.n_keys_after if row.n_keys_after is not None else "",
                    row.key_tuple,
                    row.column_name,
                    row.old_value,
                    row.new_value,
                    row.notes,
                ]
            )


def _write_change_log_excel(
    path: Path,
    summary_rows: list[dict],
    conflicts: list[dict],
    table_reviews: dict[str, TableReviewModel] | None = None,
) -> None:
    """Write slim workbook: Summary + Conflicts + review sheets (no Detail tab)."""
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    _write_header(ws, SUMMARY_HEADERS)
    for i, row in enumerate(summary_rows, start=2):
        for j, key in enumerate(SUMMARY_HEADERS, start=1):
            ws.cell(i, j, row.get(key, ""))

    ws_conf = wb.create_sheet("Conflicts")
    _write_header(ws_conf, CONFLICT_HEADERS)
    for i, c in enumerate(conflicts, start=2):
        values = [
            c.get("conflict_type", ""),
            c.get("table_name", ""),
            c.get("key_tuple", ""),
            c.get("column_name", ""),
            ", ".join(c.get("change_request_ids", [])),
            ", ".join(c.get("change_types", [])),
            ", ".join(c.get("old_values", [])),
            ", ".join(c.get("new_values", [])),
            c.get("notes", ""),
            c.get("resolved", "N"),
        ]
        for j, val in enumerate(values, start=1):
            ws_conf.cell(i, j, val)

    if table_reviews:
        write_review_sheets(wb, table_reviews)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_header(ws, headers: list[str]) -> None:
    bold = Font(bold=True)
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(1, j, h)
        cell.font = bold
