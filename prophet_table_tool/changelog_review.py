"""Human-friendly per-table review sheets for the Stage 1 Change Log."""

from __future__ import annotations

import re
from dataclasses import dataclass, field

from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .diff import ChangeRow, format_key_tuple
from .prophet_csv import ProphetTable

STRUCTURAL_TYPES = frozenset(
    {
        "column_add",
        "column_delete",
        "column_rename",
        "key_count_change",
        "table_add",
    }
)

RESERVED_SHEET_NAMES = frozenset({"Summary", "ChangeLog_Detail", "Conflicts"})

_CHANGE_COL = "_change"

_FILL_UPDATE = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_FILL_ADD = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_DELETE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_NOTE = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


@dataclass
class TableContribution:
    change_request_id: str
    before: ProphetTable | None
    after: ProphetTable | None
    change_rows: list[ChangeRow]


@dataclass
class TableReviewModel:
    table_name: str
    contributions: list[TableContribution] = field(default_factory=list)


@dataclass
class _CellAnnotation:
    old_value: str
    new_value: str
    change_request_id: str


@dataclass
class _ReviewRow:
    key_str: str
    status: str  # "", "ADD", "DELETE"
    values: dict[str, str]
    cell_updates: dict[str, list[_CellAnnotation]] = field(default_factory=dict)


@dataclass
class ReviewGrid:
    table_name: str
    columns: list[str]
    structural_notes: list[str]
    rows: list[_ReviewRow]
    multi_cr: bool


def add_contribution(
    reviews: dict[str, TableReviewModel],
    table_name: str,
    change_request_id: str,
    before: ProphetTable | None,
    after: ProphetTable | None,
    change_rows: list[ChangeRow],
) -> None:
    """Accumulate one CR's before/after snapshot for a table into the review map."""
    table_rows = [r for r in change_rows if r.table_name == table_name]
    if not table_rows and before is None and after is None:
        return
    if table_name not in reviews:
        reviews[table_name] = TableReviewModel(table_name=table_name)
    reviews[table_name].contributions.append(
        TableContribution(
            change_request_id=change_request_id,
            before=before,
            after=after,
            change_rows=table_rows,
        )
    )


def build_review_grid(model: TableReviewModel) -> ReviewGrid | None:
    """Build a wide Prophet-shaped review grid from accumulated CR contributions."""
    if not model.contributions:
        return None

    cr_ids = {c.change_request_id for c in model.contributions}
    multi_cr = len(cr_ids) > 1

    columns = _resolve_columns(model)
    if not columns:
        return None

    structural_notes = _structural_notes(model)
    row_map = _seed_rows(model, columns)
    _apply_change_annotations(model, row_map, columns)

    # Stable order: ADD rows, then unchanged/updated, then DELETE
    status_rank = {"": 1, "ADD": 0, "DELETE": 2}
    rows = sorted(
        row_map.values(),
        key=lambda r: (status_rank.get(r.status, 1), r.key_str),
    )
    return ReviewGrid(
        table_name=model.table_name,
        columns=columns,
        structural_notes=structural_notes,
        rows=rows,
        multi_cr=multi_cr,
    )


def write_review_sheets(
    wb: Workbook,
    reviews: dict[str, TableReviewModel],
) -> None:
    """Append one review sheet per table to the Change Log workbook."""
    used_names: set[str] = set(wb.sheetnames) | set(RESERVED_SHEET_NAMES)
    for table_name in sorted(reviews):
        grid = build_review_grid(reviews[table_name])
        if grid is None:
            continue
        sheet_name = _unique_sheet_name(table_name, used_names)
        used_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        _write_review_sheet(ws, grid)


def _resolve_columns(model: TableReviewModel) -> list[str]:
    columns: list[str] = []
    for contrib in model.contributions:
        src = contrib.after or contrib.before
        if src is None:
            continue
        for col in src.columns:
            if col not in columns:
                columns.append(col)
    # Also pick up columns mentioned only in change rows (e.g. column_add with empty table)
    for contrib in model.contributions:
        for row in contrib.change_rows:
            if row.change_type in {"column_add", "value_update", "row_add", "row_delete"}:
                if row.column_name and row.column_name not in columns:
                    if "->" not in row.column_name:
                        columns.append(row.column_name)
    return columns


def _structural_notes(model: TableReviewModel) -> list[str]:
    notes: list[str] = []
    seen: set[str] = set()
    for contrib in model.contributions:
        cr = contrib.change_request_id
        for row in contrib.change_rows:
            if row.change_type not in STRUCTURAL_TYPES:
                continue
            text = _format_structural_note(row, cr)
            if text not in seen:
                seen.add(text)
                notes.append(text)
    return notes


def _format_structural_note(row: ChangeRow, cr_id: str) -> str:
    prefix = f"[{cr_id}] "
    if row.change_type == "table_add":
        return f"{prefix}table_add - {row.notes or 'Table only present in after/'}"
    if row.change_type == "key_count_change":
        return (
            f"{prefix}key_count_change - {row.old_value} -> {row.new_value}"
            + (f" ({row.notes})" if row.notes else "")
        )
    if row.change_type == "column_rename":
        return f"{prefix}column_rename - {row.old_value} -> {row.new_value}"
    if row.change_type == "column_add":
        return f"{prefix}column_add - {row.column_name}"
    if row.change_type == "column_delete":
        return f"{prefix}column_delete - {row.column_name}"
    return f"{prefix}{row.change_type} - {row.column_name or row.notes}"


def _table_row_dict(table: ProphetTable) -> dict[str, dict[str, str]]:
    """Map key_str → {col: value} for all rows in a Prophet table."""
    keyed = table.with_key_tuple()
    out: dict[str, dict[str, str]] = {}
    for row in keyed.iter_rows(named=True):
        key_str = format_key_tuple(row["_key_tuple"])
        values = {
            col: "" if row[col] is None else str(row[col]) for col in table.columns
        }
        out[key_str] = values
    return out


def _seed_rows(
    model: TableReviewModel,
    columns: list[str],
) -> dict[str, _ReviewRow]:
    """Seed review rows from after snapshots, then fill gaps from before."""
    row_map: dict[str, _ReviewRow] = {}

    # Prefer after values (last contribution wins for unchanged display values)
    for contrib in model.contributions:
        if contrib.after is None:
            continue
        for key_str, values in _table_row_dict(contrib.after).items():
            if key_str not in row_map:
                row_map[key_str] = _ReviewRow(
                    key_str=key_str,
                    status="",
                    values={c: values.get(c, "") for c in columns},
                )
            else:
                for c in columns:
                    if c in values:
                        row_map[key_str].values[c] = values[c]

    # Before-only keys (candidates for DELETE) — add if missing
    for contrib in model.contributions:
        if contrib.before is None:
            continue
        for key_str, values in _table_row_dict(contrib.before).items():
            if key_str not in row_map:
                row_map[key_str] = _ReviewRow(
                    key_str=key_str,
                    status="",
                    values={c: values.get(c, "") for c in columns},
                )

    return row_map


def _apply_change_annotations(
    model: TableReviewModel,
    row_map: dict[str, _ReviewRow],
    columns: list[str],
) -> None:
    for contrib in model.contributions:
        cr = contrib.change_request_id
        add_keys: set[str] = set()
        delete_keys: set[str] = set()

        for row in contrib.change_rows:
            if row.change_type == "row_add" and row.key_tuple:
                add_keys.add(row.key_tuple)
            elif row.change_type == "row_delete" and row.key_tuple:
                delete_keys.add(row.key_tuple)

        for key_str in add_keys:
            review_row = row_map.get(key_str)
            if review_row is None:
                # Reconstruct from row_add change cells
                values = {c: "" for c in columns}
                for r in contrib.change_rows:
                    if (
                        r.change_type == "row_add"
                        and r.key_tuple == key_str
                        and r.column_name in values
                    ):
                        values[r.column_name] = r.new_value
                review_row = _ReviewRow(key_str=key_str, status="ADD", values=values)
                row_map[key_str] = review_row
            else:
                if review_row.status != "DELETE":
                    review_row.status = "ADD"

            # Fill values from row_add cells
            for r in contrib.change_rows:
                if (
                    r.change_type == "row_add"
                    and r.key_tuple == key_str
                    and r.column_name in review_row.values
                ):
                    review_row.values[r.column_name] = r.new_value

        for key_str in delete_keys:
            review_row = row_map.get(key_str)
            if review_row is None:
                values = {c: "" for c in columns}
                for r in contrib.change_rows:
                    if (
                        r.change_type == "row_delete"
                        and r.key_tuple == key_str
                        and r.column_name in values
                    ):
                        values[r.column_name] = r.old_value
                review_row = _ReviewRow(
                    key_str=key_str, status="DELETE", values=values
                )
                row_map[key_str] = review_row
            else:
                review_row.status = "DELETE"
                for r in contrib.change_rows:
                    if (
                        r.change_type == "row_delete"
                        and r.key_tuple == key_str
                        and r.column_name in review_row.values
                    ):
                        review_row.values[r.column_name] = r.old_value

        for row in contrib.change_rows:
            if row.change_type != "value_update":
                continue
            if not row.key_tuple or not row.column_name:
                continue
            if row.column_name not in columns:
                continue
            review_row = row_map.get(row.key_tuple)
            if review_row is None:
                review_row = _ReviewRow(
                    key_str=row.key_tuple,
                    status="",
                    values={c: "" for c in columns},
                )
                row_map[row.key_tuple] = review_row
            review_row.cell_updates.setdefault(row.column_name, []).append(
                _CellAnnotation(
                    old_value=row.old_value,
                    new_value=row.new_value,
                    change_request_id=cr,
                )
            )
            # Keep display new value as base when not overwritten by annotation format
            review_row.values[row.column_name] = row.new_value


def _display_cell(review_row: _ReviewRow, column: str, multi_cr: bool) -> str:
    updates = review_row.cell_updates.get(column)
    if not updates:
        return review_row.values.get(column, "")
    parts: list[str] = []
    for u in updates:
        text = f"{u.old_value} -> {u.new_value}"
        if multi_cr:
            text = f"{text} [{u.change_request_id}]"
        parts.append(text)
    return " | ".join(parts)


def _sanitize_sheet_name(name: str) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", name).strip() or "Table"
    # Excel sheet names cannot start/end with apostrophe in a problematic way;
    # also cap at 31 chars.
    return cleaned[:31]


def _unique_sheet_name(table_name: str, used: set[str]) -> str:
    base = _sanitize_sheet_name(table_name)
    if base not in used and base not in RESERVED_SHEET_NAMES:
        return base
    # Truncate to leave room for suffix _N
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        if candidate not in used and candidate not in RESERVED_SHEET_NAMES:
            return candidate
    raise ValueError(f"Unable to allocate unique sheet name for table {table_name!r}")


def _write_review_sheet(ws: Worksheet, grid: ReviewGrid) -> None:
    bold = Font(bold=True)
    row_idx = 1

    # Title / legend
    ws.cell(
        row_idx,
        1,
        f"Review: {grid.table_name} (human review only - not used by apply)",
    )
    ws.cell(row_idx, 1).font = bold
    row_idx += 1

    if grid.structural_notes:
        ws.cell(row_idx, 1, "Structural changes:")
        ws.cell(row_idx, 1).font = bold
        row_idx += 1
        for note in grid.structural_notes:
            cell = ws.cell(row_idx, 1, note)
            cell.fill = _FILL_NOTE
            row_idx += 1

    # Blank spacer when we had notes or title
    row_idx += 1

    headers = [_CHANGE_COL, *grid.columns]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row_idx, j, h)
        cell.font = bold
        cell.fill = _FILL_HEADER
    header_row = row_idx
    row_idx += 1

    for review_row in grid.rows:
        status = review_row.status
        status_cell = ws.cell(row_idx, 1, status)
        if status == "ADD":
            status_cell.fill = _FILL_ADD
        elif status == "DELETE":
            status_cell.fill = _FILL_DELETE

        for j, col in enumerate(grid.columns, start=2):
            has_update = col in review_row.cell_updates
            value = _display_cell(review_row, col, grid.multi_cr)
            cell = ws.cell(row_idx, j, value)
            if status == "ADD":
                cell.fill = _FILL_ADD
            elif status == "DELETE":
                cell.fill = _FILL_DELETE
            elif has_update:
                cell.fill = _FILL_UPDATE

        row_idx += 1

    # Light freeze: freeze below header
    ws.freeze_panes = ws.cell(header_row + 1, 2)
