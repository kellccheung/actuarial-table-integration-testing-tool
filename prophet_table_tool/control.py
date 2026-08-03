"""Control.xlsx reader — single source of truth for configuration."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from .utils import resolve_path, yn_to_bool


@dataclass
class ChangeRequestRow:
    change_request_id: str
    order: int
    include: bool
    approved: bool
    description: str = ""
    notes: str = ""


@dataclass
class ColumnRename:
    change_request_id: str
    table_name: str
    old_column_name: str
    new_column_name: str


@dataclass
class KeyCountApproval:
    change_request_id: str
    table_name: str
    approved: bool


@dataclass
class ControlConfig:
    control_path: Path
    working_root: Path
    mode: str
    production_tables_path: Path
    output_path: Path
    run_id: str
    change_requests: list[ChangeRequestRow] = field(default_factory=list)
    column_renames: list[ColumnRename] = field(default_factory=list)
    key_count_approvals: list[KeyCountApproval] = field(default_factory=list)

    def included_requests(self, *, require_approved: bool = False) -> list[ChangeRequestRow]:
        """Return included (and optionally approved) CRs sorted by order."""
        rows = [r for r in self.change_requests if r.include]
        if require_approved:
            rows = [r for r in rows if r.approved]
        return sorted(rows, key=lambda r: r.order)

    def change_request_dir(self, change_request_id: str) -> Path:
        return self.working_root / "ChangeRequests" / change_request_id

    def renames_for(self, change_request_id: str, table_name: str) -> list[ColumnRename]:
        return [
            r
            for r in self.column_renames
            if r.change_request_id == change_request_id and r.table_name == table_name
        ]

    def is_key_count_approved(self, change_request_id: str, table_name: str) -> bool:
        """True if KeyCountApprovals says Y, or the CR itself is approved."""
        for row in self.key_count_approvals:
            if (
                row.change_request_id == change_request_id
                and row.table_name == table_name
            ):
                return row.approved
        # Fall back to CR-level approved flag
        for cr in self.change_requests:
            if cr.change_request_id == change_request_id:
                return cr.approved
        return False


def read_control(control_path: Path) -> ControlConfig:
    """Load and validate Control.xlsx into a ControlConfig."""
    control_path = Path(control_path).resolve()
    wb = load_workbook(control_path, data_only=True, read_only=True)

    config_map = _read_config_sheet(wb)
    working_root_raw = config_map.get("working_root")
    if not working_root_raw:
        # Default: parent of Control.xlsx
        working_root = control_path.parent
    else:
        wr = Path(str(working_root_raw))
        working_root = wr if wr.is_absolute() else (control_path.parent / wr).resolve()

    mode = str(config_map.get("mode", "generate_changelog")).strip()
    prod_rel = str(config_map.get("production_tables_path", "Production_Tables"))
    out_rel = str(config_map.get("output_path", "Output"))
    run_id = str(config_map.get("run_id") or "").strip()

    change_requests = _read_change_requests(wb)
    column_renames = _read_column_renames(wb)
    key_count_approvals = _read_key_count_approvals(wb)

    wb.close()

    return ControlConfig(
        control_path=control_path,
        working_root=working_root,
        mode=mode,
        production_tables_path=resolve_path(working_root, prod_rel),
        output_path=resolve_path(working_root, out_rel),
        run_id=run_id,
        change_requests=change_requests,
        column_renames=column_renames,
        key_count_approvals=key_count_approvals,
    )


def _read_config_sheet(wb) -> dict[str, object]:
    if "Config" not in wb.sheetnames:
        raise ValueError("Control.xlsx missing required sheet: Config")
    ws = wb["Config"]
    result: dict[str, object] = {}
    rows = ws.iter_rows(values_only=True)
    # Expect Key | Value (with optional header)
    first = next(rows, None)
    if first is None:
        return result
    # Skip header if it looks like one
    start_rows = []
    if first[0] is not None and str(first[0]).strip().lower() in {"key", "name", "parameter"}:
        pass  # header consumed
    else:
        start_rows.append(first)

    for row in list(start_rows) + list(rows):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        value = row[1] if len(row) > 1 else None
        result[key] = value
    return result


def _read_change_requests(wb) -> list[ChangeRequestRow]:
    if "ChangeRequests" not in wb.sheetnames:
        raise ValueError("Control.xlsx missing required sheet: ChangeRequests")
    ws = wb["ChangeRequests"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return []
    col_index = _header_index(header)

    result: list[ChangeRequestRow] = []
    for row in rows_iter:
        if not row or row[0] is None:
            continue
        cr_id = str(_cell(row, col_index, "change_request_id", 0)).strip()
        if not cr_id:
            continue
        order_raw = _cell(row, col_index, "order", 1)
        try:
            order = int(order_raw) if order_raw is not None else 0
        except (TypeError, ValueError):
            order = 0
        result.append(
            ChangeRequestRow(
                change_request_id=cr_id,
                order=order,
                include=yn_to_bool(_cell(row, col_index, "include", 2)),
                approved=yn_to_bool(_cell(row, col_index, "approved", 3)),
                description=str(_cell(row, col_index, "description", 4) or ""),
                notes=str(_cell(row, col_index, "notes", 5) or ""),
            )
        )
    return result


def _read_column_renames(wb) -> list[ColumnRename]:
    if "ColumnRenames" not in wb.sheetnames:
        return []
    ws = wb["ColumnRenames"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return []
    col_index = _header_index(header)
    result: list[ColumnRename] = []
    for row in rows_iter:
        if not row or row[0] is None:
            continue
        cr_id = str(_cell(row, col_index, "change_request_id", 0) or "").strip()
        table = str(_cell(row, col_index, "table_name", 1) or "").strip()
        old_c = str(_cell(row, col_index, "old_column_name", 2) or "").strip()
        new_c = str(_cell(row, col_index, "new_column_name", 3) or "").strip()
        if cr_id and table and old_c and new_c:
            result.append(
                ColumnRename(
                    change_request_id=cr_id,
                    table_name=table,
                    old_column_name=old_c,
                    new_column_name=new_c,
                )
            )
    return result


def _read_key_count_approvals(wb) -> list[KeyCountApproval]:
    if "KeyCountApprovals" not in wb.sheetnames:
        return []
    ws = wb["KeyCountApprovals"]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        return []
    col_index = _header_index(header)
    result: list[KeyCountApproval] = []
    for row in rows_iter:
        if not row or row[0] is None:
            continue
        cr_id = str(_cell(row, col_index, "change_request_id", 0) or "").strip()
        table = str(_cell(row, col_index, "table_name", 1) or "").strip()
        approved = yn_to_bool(_cell(row, col_index, "approved", 2))
        if cr_id and table:
            result.append(
                KeyCountApproval(
                    change_request_id=cr_id,
                    table_name=table,
                    approved=approved,
                )
            )
    return result


def _header_index(header: tuple) -> dict[str, int]:
    return {
        str(h).strip().lower(): i
        for i, h in enumerate(header)
        if h is not None
    }


def _cell(row: tuple, col_index: dict[str, int], name: str, default_pos: int):
    idx = col_index.get(name.lower(), default_pos)
    if idx is None or idx >= len(row):
        return None
    return row[idx]
