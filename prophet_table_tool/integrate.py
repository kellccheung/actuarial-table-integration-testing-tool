"""Stage 2 – Validate & Integrate Change Log onto production tables."""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path
from typing import Literal

import polars as pl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .audit import close_audit_logger, setup_audit_logger
from .control import ControlConfig, read_control
from .prophet_csv import (
    ProphetTable,
    discover_csv_tables,
    normalized_key_expr,
    read_prophet_csv,
    write_prophet_csv,
)
from .utils import file_hash, generate_run_id, yn_to_bool


REPORT_HEADERS = [
    "section",
    "change_request_id",
    "table_name",
    "change_type",
    "key_tuple",
    "column_name",
    "message",
    "severity",
]


def integrate_changes(
    control_path: Path,
    change_log_path: Path,
    mode: Literal["validate_only", "apply"],
) -> Path:
    """
    Stage 2: Validate (and optionally apply) a Change Log onto production tables.

    Returns the path to IntegrationReport_*.xlsx.
    """
    if mode not in {"validate_only", "apply"}:
        raise ValueError(f"mode must be 'validate_only' or 'apply', got {mode!r}")

    control_path = Path(control_path)
    change_log_path = Path(change_log_path)
    control = read_control(control_path)
    run_id = control.run_id or generate_run_id()
    control.output_path.mkdir(parents=True, exist_ok=True)

    logger, _log_path = setup_audit_logger(control.output_path, run_id, mode)
    status = "FAILED"
    report_path = control.output_path / f"IntegrationReport_{run_id}.xlsx"
    validation_messages: list[dict] = []

    try:
        logger.info("control_path=%s", control_path)
        logger.info("control_hash=%s", file_hash(control_path))
        logger.info("change_log_path=%s", change_log_path)
        logger.info("change_log_hash=%s", file_hash(change_log_path))
        logger.info("production_tables_path=%s", control.production_tables_path)

        approved_crs = control.included_requests(require_approved=True)
        cr_order = {r.change_request_id: r.order for r in approved_crs}
        cr_ids = {r.change_request_id for r in approved_crs}
        logger.info("change_requests_processed=%s", [r.change_request_id for r in approved_crs])

        detail_rows, conflicts = _load_change_log(change_log_path)
        detail_rows = [r for r in detail_rows if r["change_request_id"] in cr_ids]
        detail_rows.sort(
            key=lambda r: (
                cr_order.get(r["change_request_id"], 10**9),
                r["table_name"],
                r["change_type"],
                r["key_tuple"],
                r["column_name"],
            )
        )
        logger.info("loaded %d detail change row(s)", len(detail_rows))

        unresolved = [
            c
            for c in conflicts
            if not yn_to_bool(c.get("resolved", "N"))
            and any(cid in cr_ids for cid in c.get("change_request_ids", []))
        ]
        if unresolved:
            for c in unresolved:
                msg = (
                    f"Unresolved conflict ({c.get('conflict_type')}) on table "
                    f"{c.get('table_name')} key={c.get('key_tuple')!r} "
                    f"col={c.get('column_name')!r} CRs={c.get('change_request_ids')}"
                )
                validation_messages.append(
                    _vmsg("conflict", "", c.get("table_name", ""), "", "", "", msg, "FAIL")
                )
                logger.error(msg)
            if mode == "apply":
                status = "FAILED"
                _write_integration_report(report_path, validation_messages, status, mode)
                logger.info("final_status=%s", status)
                return report_path

        logger.info("loading production tables from %s", control.production_tables_path)
        prod_paths = discover_csv_tables(control.production_tables_path)
        logger.info("discovered %d production table(s)", len(prod_paths))
        tables: dict[str, ProphetTable] = {}
        for name, path in prod_paths.items():
            logger.info("reading production table %s", name)
            tables[name] = read_prophet_csv(path)

        by_cr_table: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
        for row in detail_rows:
            by_cr_table[row["change_request_id"]][row["table_name"]].append(row)

        ok = True
        for cr in approved_crs:
            for table_name, rows in by_cr_table.get(cr.change_request_id, {}).items():
                logger.info(
                    "[%s] validating %s (%d change row(s))...",
                    cr.change_request_id,
                    table_name,
                    len(rows),
                )
                table_ok, msgs = _validate_table_changes(
                    control, cr.change_request_id, table_name, rows, tables
                )
                validation_messages.extend(msgs)
                if not table_ok:
                    ok = False
                    logger.warning("[%s] validation FAILED for %s", cr.change_request_id, table_name)
                else:
                    logger.info("[%s] validation OK for %s", cr.change_request_id, table_name)

        if not ok:
            status = "FAILED"
            _write_integration_report(report_path, validation_messages, status, mode)
            logger.info(
                "n_validation_failures=%d",
                sum(1 for m in validation_messages if m["severity"] == "FAIL"),
            )
            logger.info("final_status=%s", status)
            return report_path

        if mode == "validate_only":
            status = "DRY_RUN_SUCCESS"
            validation_messages.append(
                _vmsg(
                    "summary",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "Validation passed (dry-run); no tables written.",
                    "INFO",
                )
            )
            _write_integration_report(report_path, validation_messages, status, mode)
            logger.info("tables_affected=%d", len({r["table_name"] for r in detail_rows}))
            logger.info("final_status=%s", status)
            return report_path

        for cr in approved_crs:
            for table_name, rows in by_cr_table.get(cr.change_request_id, {}).items():
                logger.info(
                    "[%s] applying %s (%d change row(s))...",
                    cr.change_request_id,
                    table_name,
                    len(rows),
                )
                tables[table_name] = _apply_table_changes(
                    control, cr.change_request_id, table_name, rows, tables
                )
                validation_messages.append(
                    _vmsg(
                        "apply",
                        cr.change_request_id,
                        table_name,
                        "",
                        "",
                        "",
                        f"Applied {len(rows)} change row(s).",
                        "INFO",
                    )
                )
                logger.info("[%s] applied %s", cr.change_request_id, table_name)

        out_dir = control.output_path / "New_Production_Tables"
        out_dir.mkdir(parents=True, exist_ok=True)

        written = 0
        for name, table in tables.items():
            out_path = out_dir / Path(name).with_suffix(table.source_suffix)
            logger.info("writing %s -> %s", name, out_path)
            write_prophet_csv(table, out_path)
            written += 1

        validation_messages.append(
            _vmsg("summary", "", "", "", "", "", f"Wrote {written} table(s) to {out_dir}", "INFO")
        )
        status = "SUCCESS"
        _write_integration_report(report_path, validation_messages, status, mode)
        logger.info("tables_affected=%d", written)
        logger.info("output_dir=%s", out_dir)
        logger.info("final_status=%s", status)
        return report_path

    except Exception as exc:
        logger.exception("integrate_changes failed: %s", exc)
        validation_messages.append(
            _vmsg("error", "", "", "", "", "", str(exc), "FAIL")
        )
        try:
            _write_integration_report(report_path, validation_messages, status, mode)
        except Exception:
            pass
        logger.info("final_status=%s", status)
        raise
    finally:
        close_audit_logger(logger)


def _vmsg(
    section: str,
    cr: str,
    table: str,
    change_type: str,
    key_tuple: str,
    column: str,
    message: str,
    severity: str,
) -> dict:
    return {
        "section": section,
        "change_request_id": cr,
        "table_name": table,
        "change_type": change_type,
        "key_tuple": key_tuple,
        "column_name": column,
        "message": message,
        "severity": severity,
    }


def _detail_csv_path_for(change_log_path: Path) -> Path:
    """Sidecar path: ChangeLog_{run_id}.xlsx → ChangeLog_{run_id}_Detail.csv."""
    return change_log_path.with_name(f"{change_log_path.stem}_Detail.csv")


def _load_change_log(path: Path) -> tuple[list[dict], list[dict]]:
    detail_csv = _detail_csv_path_for(path)
    if detail_csv.is_file():
        detail = _load_detail_csv(detail_csv)
    else:
        detail = _load_detail_from_excel(path)

    conflicts: list[dict] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Conflicts" in wb.sheetnames:
        ws = wb["Conflicts"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            keys = [str(h).strip() if h else f"col{i}" for i, h in enumerate(header)]
            for row in rows:
                if not row or all(c is None for c in row):
                    continue
                d = {keys[i]: (row[i] if i < len(row) else None) for i in range(len(keys))}
                cr_raw = str(d.get("change_request_ids") or "")
                conflicts.append(
                    {
                        "conflict_type": str(d.get("conflict_type") or ""),
                        "table_name": str(d.get("table_name") or ""),
                        "key_tuple": str(d.get("key_tuple") or ""),
                        "column_name": str(d.get("column_name") or ""),
                        "change_request_ids": [x.strip() for x in cr_raw.split(",") if x.strip()],
                        "resolved": d.get("resolved", "N"),
                        "notes": str(d.get("notes") or ""),
                    }
                )
    wb.close()
    return detail, conflicts


def _load_detail_csv(path: Path) -> list[dict]:
    detail: list[dict] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for d in reader:
            if not d or not (d.get("change_request_id") or "").strip():
                continue
            detail.append(
                {
                    "change_request_id": str(d.get("change_request_id") or "").strip(),
                    "table_name": str(d.get("table_name") or "").strip(),
                    "change_type": str(d.get("change_type") or "").strip(),
                    "n_keys_before": d.get("n_keys_before") or None,
                    "n_keys_after": d.get("n_keys_after") or None,
                    "key_tuple": str(d.get("key_tuple") or ""),
                    "column_name": str(d.get("column_name") or ""),
                    "old_value": "" if d.get("old_value") is None else str(d.get("old_value")),
                    "new_value": "" if d.get("new_value") is None else str(d.get("new_value")),
                    "notes": str(d.get("notes") or ""),
                }
            )
    return detail


def _load_detail_from_excel(path: Path) -> list[dict]:
    """Fallback for older Change Logs that still embed ChangeLog_Detail."""
    detail: list[dict] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    if "ChangeLog_Detail" in wb.sheetnames:
        ws = wb["ChangeLog_Detail"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header:
            keys = [str(h).strip() if h else f"col{i}" for i, h in enumerate(header)]
            for row in rows:
                if not row or row[0] is None:
                    continue
                d = {keys[i]: (row[i] if i < len(row) else None) for i in range(len(keys))}
                detail.append(
                    {
                        "change_request_id": str(d.get("change_request_id") or "").strip(),
                        "table_name": str(d.get("table_name") or "").strip(),
                        "change_type": str(d.get("change_type") or "").strip(),
                        "n_keys_before": d.get("n_keys_before"),
                        "n_keys_after": d.get("n_keys_after"),
                        "key_tuple": str(d.get("key_tuple") or ""),
                        "column_name": str(d.get("column_name") or ""),
                        "old_value": "" if d.get("old_value") is None else str(d.get("old_value")),
                        "new_value": "" if d.get("new_value") is None else str(d.get("new_value")),
                        "notes": str(d.get("notes") or ""),
                    }
                )
    wb.close()
    return detail


def _validate_table_changes(
    control: ControlConfig,
    cr_id: str,
    table_name: str,
    rows: list[dict],
    tables: dict[str, ProphetTable],
) -> tuple[bool, list[dict]]:
    msgs: list[dict] = []
    ok = True
    types = {r["change_type"] for r in rows}
    is_table_add = "table_add" in types

    if not is_table_add and table_name not in tables:
        ok = False
        msgs.append(
            _vmsg(
                "validation",
                cr_id,
                table_name,
                "",
                "",
                "",
                f"Table '{table_name}' not found in production (and not a table_add).",
                "FAIL",
            )
        )
        return ok, msgs

    for r in rows:
        if r["change_type"] == "key_count_change":
            if not control.is_key_count_approved(cr_id, table_name):
                ok = False
                msgs.append(
                    _vmsg(
                        "validation",
                        cr_id,
                        table_name,
                        "key_count_change",
                        "",
                        "",
                        "Key-count change blocked: not approved in Control.",
                        "FAIL",
                    )
                )

    for r in rows:
        if r["change_type"] == "column_rename":
            declared = control.renames_for(cr_id, table_name)
            old_c = r["old_value"]
            new_c = r["new_value"]
            if not any(
                d.old_column_name == old_c and d.new_column_name == new_c for d in declared
            ):
                ok = False
                msgs.append(
                    _vmsg(
                        "validation",
                        cr_id,
                        table_name,
                        "column_rename",
                        "",
                        r["column_name"],
                        f"Column rename {old_c!r}->{new_c!r} not declared in ColumnRenames.",
                        "FAIL",
                    )
                )

    if is_table_add or table_name not in tables:
        return ok, msgs

    table = tables[table_name]
    keyed = table.with_key_tuple()

    lookup: dict[str, dict] = {}
    for row in keyed.iter_rows(named=True):
        key_str = row["_key_str"]
        lookup[key_str] = row

    for r in rows:
        ct = r["change_type"]
        if ct not in {"value_update", "row_delete"}:
            continue
        key_str = r["key_tuple"]
        col = r["column_name"]
        if key_str not in lookup:
            ok = False
            msgs.append(
                _vmsg(
                    "validation",
                    cr_id,
                    table_name,
                    ct,
                    key_str,
                    col,
                    f"Key not found in production for {ct}.",
                    "FAIL",
                )
            )
            continue
        # old_value in the Change Log is informational only — it is not required
        # to match the current production cell.

    if ok:
        msgs.append(
            _vmsg(
                "validation",
                cr_id,
                table_name,
                "",
                "",
                "",
                "Validation passed.",
                "INFO",
            )
        )
    return ok, msgs


def _apply_table_changes(
    control: ControlConfig,
    cr_id: str,
    table_name: str,
    rows: list[dict],
    tables: dict[str, ProphetTable],
) -> ProphetTable:
    """Apply all change rows for one CR+table; return the updated ProphetTable."""
    types = {r["change_type"] for r in rows}

    if "table_add" in types:
        n_keys = None
        for r in rows:
            if r["n_keys_after"] is not None and str(r["n_keys_after"]).strip() != "":
                n_keys = int(r["n_keys_after"])
                break
        if n_keys is None:
            n_keys = 1

        add_rows = [r for r in rows if r["change_type"] == "row_add"]
        col_order: list[str] = []
        by_key: dict[str, dict[str, str]] = {}
        for r in add_rows:
            col = r["column_name"]
            if col and col not in col_order:
                col_order.append(col)
            by_key.setdefault(r["key_tuple"], {})[col] = r["new_value"]

        records = []
        for key in sorted(by_key.keys()):
            vals = by_key[key]
            records.append([vals.get(c, "") for c in col_order])
        if records:
            data = pl.DataFrame(records, schema=col_order, orient="row").cast(pl.Utf8)
        else:
            data = pl.DataFrame({c: [] for c in col_order}).cast(pl.Utf8)

        return ProphetTable(n_keys=n_keys, columns=col_order, data=data)

    table = tables[table_name]

    original_n_keys = table.n_keys
    n_keys = table.n_keys
    columns = list(table.columns)
    data = table.data
    leading_dummy_lines = list(table.leading_dummy_lines)
    trailing_dummy_lines = list(table.trailing_dummy_lines)
    source_path = table.source_path

    for r in rows:
        if r["change_type"] == "key_count_change" and r["n_keys_after"] is not None:
            if str(r["n_keys_after"]).strip() != "":
                n_keys = int(r["n_keys_after"])

    for r in rows:
        if r["change_type"] == "column_rename":
            old_c, new_c = r["old_value"], r["new_value"]
            if old_c in columns:
                columns = [new_c if c == old_c else c for c in columns]
                data = data.rename({old_c: new_c})

    for r in rows:
        if r["change_type"] == "column_add":
            col = r["column_name"]
            if col and col not in columns:
                if n_keys > original_n_keys:
                    insert_at = max(original_n_keys - 1, 0)
                    insert_at = min(insert_at, len(columns))
                    columns.insert(insert_at, col)
                else:
                    columns.append(col)
                data = data.with_columns(pl.lit("").cast(pl.Utf8).alias(col))
                data = data.select(columns)

    for r in rows:
        if r["change_type"] == "column_delete":
            col = r["column_name"]
            if col in columns:
                columns = [c for c in columns if c != col]
                data = data.drop(col)

    match_key_cols = columns[: max(original_n_keys - 1, 0)]
    final_key_cols = columns[: max(n_keys - 1, 0)]

    def key_str_expr(key_cols: list[str]):
        if not key_cols:
            return pl.lit("")
        return pl.concat_str(
            [normalized_key_expr(c) for c in key_cols],
            separator="|",
        )

    data = data.with_columns(key_str_expr(match_key_cols).alias("_key_str"))

    delete_keys = {r["key_tuple"] for r in rows if r["change_type"] == "row_delete"}
    if delete_keys:
        data = data.filter(~pl.col("_key_str").is_in(list(delete_keys)))

    updates = [r for r in rows if r["change_type"] == "value_update"]
    if updates:
        by_key: dict[str, dict[str, str]] = defaultdict(dict)
        for r in updates:
            if r["column_name"]:
                by_key[r["key_tuple"]][r["column_name"]] = r["new_value"]

        for key_str, col_vals in by_key.items():
            mask = pl.col("_key_str") == key_str
            for col, new_val in col_vals.items():
                if col not in data.columns:
                    data = data.with_columns(pl.lit("").cast(pl.Utf8).alias(col))
                    if col not in columns:
                        columns.append(col)
                data = data.with_columns(
                    pl.when(mask)
                    .then(pl.lit(new_val).cast(pl.Utf8))
                    .otherwise(pl.col(col))
                    .alias(col)
                )

    add_rows = [r for r in rows if r["change_type"] == "row_add"]
    if add_rows:
        by_key = defaultdict(dict)
        for r in add_rows:
            if r["column_name"]:
                by_key[r["key_tuple"]][r["column_name"]] = r["new_value"]
        new_records = []
        for key_str, col_vals in sorted(by_key.items()):
            key_parts = key_str.split("|") if key_str else []
            record = {c: col_vals.get(c, "") for c in columns}
            apply_keys = (
                final_key_cols
                if len(key_parts) == len(final_key_cols)
                else match_key_cols
            )
            for i, kc in enumerate(apply_keys):
                if i < len(key_parts):
                    record[kc] = key_parts[i]
            for c, v in col_vals.items():
                record[c] = v
                if c not in columns:
                    columns.append(c)
            new_records.append([record.get(c, "") for c in columns])

        for c in columns:
            if c not in data.columns and c != "_key_str":
                data = data.with_columns(pl.lit("").cast(pl.Utf8).alias(c))

        if new_records:
            new_df = pl.DataFrame(new_records, schema=columns, orient="row").cast(pl.Utf8)
            new_df = new_df.with_columns(key_str_expr(match_key_cols).alias("_key_str"))
            data = pl.concat(
                [
                    data.select(columns + ["_key_str"]),
                    new_df.select(columns + ["_key_str"]),
                ],
                how="vertical",
            )

    data = data.select(columns)
    return ProphetTable(
        n_keys=n_keys,
        columns=columns,
        data=data,
        source_path=source_path,
        leading_dummy_lines=leading_dummy_lines,
        trailing_dummy_lines=trailing_dummy_lines,
    )


def _write_integration_report(
    path: Path,
    messages: list[dict],
    status: str,
    mode: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation_Report"
    bold = Font(bold=True)
    headers = REPORT_HEADERS + ["status", "mode"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(1, j, h)
        cell.font = bold

    for i, m in enumerate(messages, start=2):
        for j, key in enumerate(REPORT_HEADERS, start=1):
            ws.cell(i, j, m.get(key, ""))
        ws.cell(i, len(REPORT_HEADERS) + 1, status)
        ws.cell(i, len(REPORT_HEADERS) + 2, mode)

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "status"
    ws2["B1"] = status
    ws2["A2"] = "mode"
    ws2["B2"] = mode
    ws2["A3"] = "n_messages"
    ws2["B3"] = len(messages)
    ws2["A4"] = "n_failures"
    ws2["B4"] = sum(1 for m in messages if m.get("severity") == "FAIL")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
