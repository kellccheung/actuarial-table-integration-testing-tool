"""
Acceptance tests T01–T12 for the Prophet Table Change Tool.

Requires fixtures under fixtures/acceptance/ (run build script first):
    python fixtures/build_acceptance_fixtures.py
    pytest tests/test_acceptance_t01_t12.py -v
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from prophet_table_tool import generate_change_log, integrate_changes

FIXTURES = Path(__file__).resolve().parents[1] / "fixtures" / "acceptance"


def _case(name: str) -> Path:
    path = FIXTURES / name
    if not path.exists():
        pytest.skip(
            f"Fixture {name} missing — run: python fixtures/build_acceptance_fixtures.py"
        )
    return path


def _clean_output(case_dir: Path) -> None:
    out = case_dir / "Output"
    if out.exists():
        shutil.rmtree(out)
    out.mkdir(parents=True)
    (out / "Audit").mkdir(parents=True)


def _sheet_rows(xlsx: Path, sheet: str) -> list[dict]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb[sheet]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    keys = [str(h) for h in header]
    result = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        result.append({keys[i]: row[i] if i < len(row) else None for i in range(len(keys))})
    wb.close()
    return result


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8").replace("\r\n", "\n").strip()


def _assert_csv_equal(actual: Path, expected: Path) -> None:
    assert actual.exists(), f"Missing output table: {actual}"
    assert _read_text(actual) == _read_text(expected), (
        f"CSV mismatch:\n--- actual {actual} ---\n{_read_text(actual)}\n"
        f"--- expected {expected} ---\n{_read_text(expected)}"
    )


# ---------------------------------------------------------------------------
# T01
# ---------------------------------------------------------------------------

def test_t01_two_crs_no_overlap():
    case = _case("T01_two_crs_no_overlap")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    detail = _sheet_rows(clog, "ChangeLog_Detail")
    conflicts = _sheet_rows(clog, "Conflicts")
    assert conflicts == [] or all(
        not r.get("conflict_type") for r in conflicts
    ) or len(conflicts) == 0
    # No conflicts sheet rows with content
    assert not any(r.get("conflict_type") for r in conflicts)

    types = {r["change_type"] for r in detail}
    assert "value_update" in types
    assert "CR_T01_Mortality" in {r["change_request_id"] for r in detail}
    assert "CR_T01_Expense" in {r["change_request_id"] for r in detail}

    # Per-table human review sheets (ignored by Stage 2)
    wb_clog = load_workbook(clog, data_only=True)
    assert "MORT_TABLE" in wb_clog.sheetnames
    assert "EXPENSE_TABLE" in wb_clog.sheetnames
    mort_rows = list(wb_clog["MORT_TABLE"].iter_rows(values_only=True))
    expense_rows = list(wb_clog["EXPENSE_TABLE"].iter_rows(values_only=True))
    wb_clog.close()
    mort_header = next(r for r in mort_rows if r and r[0] == "_change")
    assert list(mort_header[1:4]) == ["Age", "Duration", "Product"]
    assert any(
        cell is not None and "0.0012 -> 0.0013" in str(cell)
        for row in mort_rows
        for cell in row
    )
    assert any(
        cell is not None and "100 -> 110" in str(cell)
        for row in expense_rows
        for cell in row
    )

    report = integrate_changes(case / "Control.xlsx", clog, "apply")
    wb = load_workbook(report, data_only=True)
    assert wb["Summary"]["B1"].value == "SUCCESS"
    wb.close()

    out = case / "Output" / "New_Production_Tables"
    _assert_csv_equal(out / "MORT_TABLE.csv", case / "expected" / "MORT_TABLE.csv")
    _assert_csv_equal(out / "EXPENSE_TABLE.csv", case / "expected" / "EXPENSE_TABLE.csv")


# ---------------------------------------------------------------------------
# T02
# ---------------------------------------------------------------------------

def test_t02_same_cell_conflict():
    case = _case("T02_same_cell_conflict")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    conflicts = _sheet_rows(clog, "Conflicts")
    assert any(r.get("conflict_type") == "cell_overlap" for r in conflicts)

    report = integrate_changes(case / "Control.xlsx", clog, "apply")
    wb = load_workbook(report, data_only=True)
    assert wb["Summary"]["B1"].value == "FAILED"
    wb.close()

    new_dir = case / "Output" / "New_Production_Tables"
    assert not new_dir.exists() or list(new_dir.glob("*.csv")) == []


# ---------------------------------------------------------------------------
# T03
# ---------------------------------------------------------------------------

def test_t03_before_only_table():
    case = _case("T03_before_only_table")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    detail = _sheet_rows(clog, "ChangeLog_Detail")
    assert not any(r.get("table_name") == "OBSOLETE_TABLE" for r in detail)
    assert any(r.get("table_name") == "MORT_TABLE" for r in detail)

    audit = list((case / "Output" / "Audit").glob("*generate_changelog.log"))
    assert audit
    text = audit[0].read_text(encoding="utf-8")
    assert "OBSOLETE_TABLE" in text
    assert "only in before" in text.lower() or "skipped" in text.lower()


# ---------------------------------------------------------------------------
# T04
# ---------------------------------------------------------------------------

def test_t04_table_add():
    case = _case("T04_after_only_table_add")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    detail = _sheet_rows(clog, "ChangeLog_Detail")
    assert any(r.get("change_type") == "table_add" for r in detail)
    assert any(r.get("table_name") == "NEW_LOADING_TABLE" for r in detail)

    integrate_changes(case / "Control.xlsx", clog, "apply")
    out = case / "Output" / "New_Production_Tables"
    _assert_csv_equal(
        out / "NEW_LOADING_TABLE.csv", case / "expected" / "NEW_LOADING_TABLE.csv"
    )
    assert (out / "MORT_TABLE.csv").exists()


# ---------------------------------------------------------------------------
# T05
# ---------------------------------------------------------------------------

def test_t05_column_rename_declared():
    case = _case("T05_column_rename_declared")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    detail = _sheet_rows(clog, "ChangeLog_Detail")
    assert any(r.get("change_type") == "column_rename" for r in detail)

    integrate_changes(case / "Control.xlsx", clog, "apply")
    out_csv = case / "Output" / "New_Production_Tables" / "MORT_TABLE.csv"
    text = _read_text(out_csv)
    assert "OldRate" not in text
    assert ",Rate," in text or text.split("\n")[0].endswith("Rate,Loading") or "Rate" in text.split("\n")[0]
    _assert_csv_equal(out_csv, case / "expected" / "MORT_TABLE.csv")


# ---------------------------------------------------------------------------
# T06
# ---------------------------------------------------------------------------

def test_t06_column_rename_undeclared():
    case = _case("T06_column_rename_undeclared")
    _clean_output(case)

    # Stage 1 with declared rename → ChangeLog contains column_rename
    clog = generate_change_log(case / "Control.xlsx")
    detail = _sheet_rows(clog, "ChangeLog_Detail")
    assert any(r.get("change_type") == "column_rename" for r in detail)

    # Stage 2 with Control_stage2 (no ColumnRenames) → hard stop
    report = integrate_changes(case / "Control_stage2.xlsx", clog, "apply")
    msgs = _sheet_rows(report, "Validation_Report")
    assert any(
        "not declared" in str(m.get("message", "")).lower() for m in msgs
    )
    wb = load_workbook(report, data_only=True)
    assert wb["Summary"]["B1"].value == "FAILED"
    wb.close()
    new_dir = case / "Output" / "New_Production_Tables"
    assert not new_dir.exists() or list(new_dir.glob("*.csv")) == []


# ---------------------------------------------------------------------------
# T07
# ---------------------------------------------------------------------------

def test_t07_key_count_approved():
    case = _case("T07_key_count_approved")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    detail = _sheet_rows(clog, "ChangeLog_Detail")
    assert any(r.get("change_type") == "key_count_change" for r in detail)

    report = integrate_changes(case / "Control.xlsx", clog, "apply")
    wb = load_workbook(report, data_only=True)
    assert wb["Summary"]["B1"].value == "SUCCESS"
    wb.close()

    out_csv = case / "Output" / "New_Production_Tables" / "MORT_TABLE.csv"
    header = out_csv.read_text(encoding="utf-8").splitlines()[0]
    assert header.startswith("!5,")
    assert "Sex" in header
    _assert_csv_equal(out_csv, case / "expected" / "MORT_TABLE.csv")


# ---------------------------------------------------------------------------
# T08
# ---------------------------------------------------------------------------

def test_t08_key_count_not_approved():
    case = _case("T08_key_count_not_approved")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    assert any(
        r.get("change_type") == "key_count_change"
        for r in _sheet_rows(clog, "ChangeLog_Detail")
    )

    report = integrate_changes(case / "Control.xlsx", clog, "apply")
    msgs = _sheet_rows(report, "Validation_Report")
    assert any("not approved" in str(m.get("message", "")).lower() for m in msgs)
    wb = load_workbook(report, data_only=True)
    assert wb["Summary"]["B1"].value == "FAILED"
    wb.close()
    new_dir = case / "Output" / "New_Production_Tables"
    assert not new_dir.exists() or list(new_dir.glob("*.csv")) == []


# ---------------------------------------------------------------------------
# T09
# ---------------------------------------------------------------------------

def test_t09_validate_only():
    case = _case("T09_validate_only")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    report = integrate_changes(case / "Control.xlsx", clog, "validate_only")
    wb = load_workbook(report, data_only=True)
    assert wb["Summary"]["B1"].value == "DRY_RUN_SUCCESS"
    assert wb["Summary"]["B2"].value == "validate_only"
    wb.close()

    new_dir = case / "Output" / "New_Production_Tables"
    assert not new_dir.exists() or list(new_dir.glob("*.csv")) == []


# ---------------------------------------------------------------------------
# T10
# ---------------------------------------------------------------------------

def test_t10_old_value_mismatch():
    case = _case("T10_old_value_mismatch")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    report = integrate_changes(case / "Control.xlsx", clog, "apply")
    msgs = _sheet_rows(report, "Validation_Report")
    assert any("old_value mismatch" in str(m.get("message", "")).lower() for m in msgs)
    assert any("20|1|PROD_A" in str(m.get("key_tuple", "")) for m in msgs)
    wb = load_workbook(report, data_only=True)
    assert wb["Summary"]["B1"].value == "FAILED"
    wb.close()


# ---------------------------------------------------------------------------
# T11
# ---------------------------------------------------------------------------

def test_t11_idempotent_apply():
    case = _case("T11_idempotent_apply")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    clog_keep = case / "ChangeLog_kept.xlsx"
    shutil.copy(clog, clog_keep)

    integrate_changes(case / "Control.xlsx", clog_keep, "apply")
    out = case / "Output" / "New_Production_Tables"
    first = {p.name: _read_text(p) for p in out.glob("*.csv")}

    # Clear outputs only; re-apply same Change Log onto the same production snapshot
    shutil.rmtree(out)
    integrate_changes(case / "Control.xlsx", clog_keep, "apply")
    second = {
        p.name: _read_text(p)
        for p in (case / "Output" / "New_Production_Tables").glob("*.csv")
    }

    assert first == second
    _assert_csv_equal(
        case / "Output" / "New_Production_Tables" / "MORT_TABLE.csv",
        case / "expected" / "MORT_TABLE.csv",
    )


# ---------------------------------------------------------------------------
# T12
# ---------------------------------------------------------------------------

def test_t12_empty_change_request():
    case = _case("T12_empty_change_request")
    _clean_output(case)

    clog = generate_change_log(case / "Control.xlsx")
    summary = _sheet_rows(clog, "Summary")
    assert len(summary) == 1
    assert summary[0]["change_request_id"] == "CR_T12_Empty"
    assert summary[0]["status"] == "EMPTY"
    detail = _sheet_rows(clog, "ChangeLog_Detail")
    assert detail == []
