"""Tests for encoding fallback, numeric compare, subfolders, and sparse reviews."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from prophet_table_tool.changelog import generate_change_log
from prophet_table_tool.control import ControlConfig
from prophet_table_tool.diff import ChangeRow, detect_conflicts, diff_table
from prophet_table_tool.prophet_csv import discover_csv_tables, read_prophet_csv
from prophet_table_tool.utils import normalize_key_part, values_equal


def _minimal_control(tmp_path: Path | None = None) -> ControlConfig:
    root = tmp_path or Path(".")
    return ControlConfig(
        control_path=root / "Control.xlsx",
        working_root=root,
        mode="generate_changelog",
        production_tables_path=root / "Production_Tables",
        output_path=root / "Output",
        run_id="TEST",
    )


def test_values_equal_numeric_formats():
    assert values_equal("1.10", "1.1")
    assert values_equal("1.10", "1.2") is False
    assert values_equal("abc", "abc")
    assert values_equal("abc", "abd") is False
    assert values_equal("", "")
    assert values_equal("", "0") is False


def test_normalize_key_part_keeps_integers():
    assert normalize_key_part("20") == "20"
    assert normalize_key_part("1.10") == "1.1"
    assert normalize_key_part("PROD_A") == "PROD_A"


def test_read_cp1252_fallback(tmp_path: Path):
    # Byte 0xA3 is £ in cp1252; invalid as UTF-8 continuation alone in some contexts.
    # Use a cp1252-only character sequence that fails utf-8.
    content = "!2,Product,Note\n*,PROD_A,caf\xe9 rate\n"
    path = tmp_path / "TABLE.csv"
    path.write_bytes(content.encode("cp1252"))

    table = read_prophet_csv(path)
    assert table.source_encoding in {"cp1252", "latin-1"}
    assert table.data.height == 1
    assert "caf" in table.data["Note"][0]


def test_numeric_value_compare_no_false_diff(tmp_path: Path):
    before_path = tmp_path / "before.csv"
    after_path = tmp_path / "after.csv"
    before_path.write_text(
        "!4,Age,Duration,Product,Rate,Loading\n"
        "*,20,1,PROD_A,1.10,1.05\n",
        encoding="utf-8",
    )
    after_path.write_text(
        "!4,Age,Duration,Product,Rate,Loading\n"
        "*,20,1,PROD_A,1.1,1.05\n",
        encoding="utf-8",
    )
    result = diff_table(
        "CR1",
        "MORT_TABLE",
        read_prophet_csv(before_path),
        read_prophet_csv(after_path),
        _minimal_control(tmp_path),
    )
    assert not any(r.change_type == "value_update" for r in result.change_rows)


def test_numeric_value_compare_detects_real_diff(tmp_path: Path):
    before_path = tmp_path / "before.csv"
    after_path = tmp_path / "after.csv"
    before_path.write_text(
        "!4,Age,Duration,Product,Rate,Loading\n"
        "*,20,1,PROD_A,1.10,1.05\n",
        encoding="utf-8",
    )
    after_path.write_text(
        "!4,Age,Duration,Product,Rate,Loading\n"
        "*,20,1,PROD_A,1.2,1.05\n",
        encoding="utf-8",
    )
    result = diff_table(
        "CR1",
        "MORT_TABLE",
        read_prophet_csv(before_path),
        read_prophet_csv(after_path),
        _minimal_control(tmp_path),
    )
    updates = [r for r in result.change_rows if r.change_type == "value_update"]
    assert len(updates) == 1
    assert updates[0].column_name == "Rate"
    assert updates[0].old_value == "1.10"
    assert updates[0].new_value == "1.2"


def test_numeric_key_joins_despite_format(tmp_path: Path):
    before_path = tmp_path / "before.csv"
    after_path = tmp_path / "after.csv"
    before_path.write_text(
        "!2,Age,Rate\n*,1.10,0.5\n",
        encoding="utf-8",
    )
    after_path.write_text(
        "!2,Age,Rate\n*,1.1,0.6\n",
        encoding="utf-8",
    )
    result = diff_table(
        "CR1",
        "T",
        read_prophet_csv(before_path),
        read_prophet_csv(after_path),
        _minimal_control(tmp_path),
    )
    assert not any(r.change_type in {"row_add", "row_delete"} for r in result.change_rows)
    updates = [r for r in result.change_rows if r.change_type == "value_update"]
    assert len(updates) == 1
    assert updates[0].key_tuple == "1.1"


def test_discover_recursive_subfolders(tmp_path: Path):
    sub_a = tmp_path / "SubA"
    sub_b = tmp_path / "SubB"
    sub_a.mkdir()
    sub_b.mkdir()
    (sub_a / "MORT_TABLE.csv").write_text(
        "!2,Age,Rate\n*,20,0.1\n", encoding="utf-8"
    )
    (sub_b / "MORT_TABLE.csv").write_text(
        "!2,Age,Rate\n*,30,0.2\n", encoding="utf-8"
    )
    (tmp_path / "FLAT.csv").write_text("!2,X,Y\n*,1,2\n", encoding="utf-8")

    found = discover_csv_tables(tmp_path)
    assert set(found) == {"SubA/MORT_TABLE", "SubB/MORT_TABLE", "FLAT"}
    assert found["SubA/MORT_TABLE"].parent.name == "SubA"


def test_subfolder_tables_compared_separately(tmp_path: Path):
    """End-to-end: same stem in different subfolders yields separate identities."""
    root = tmp_path
    cr = "CR_SUB"
    before = root / "ChangeRequests" / cr / "before"
    after = root / "ChangeRequests" / cr / "after"
    (before / "SubA").mkdir(parents=True)
    (before / "SubB").mkdir(parents=True)
    (after / "SubA").mkdir(parents=True)
    (after / "SubB").mkdir(parents=True)

    (before / "SubA" / "T.csv").write_text("!2,Age,Rate\n*,20,1.0\n", encoding="utf-8")
    (after / "SubA" / "T.csv").write_text("!2,Age,Rate\n*,20,1.1\n", encoding="utf-8")
    (before / "SubB" / "T.csv").write_text("!2,Age,Rate\n*,30,2.0\n", encoding="utf-8")
    (after / "SubB" / "T.csv").write_text("!2,Age,Rate\n*,30,2.0\n", encoding="utf-8")

    prod = root / "Production_Tables"
    (prod / "SubA").mkdir(parents=True)
    (prod / "SubB").mkdir(parents=True)
    (prod / "SubA" / "T.csv").write_text("!2,Age,Rate\n*,20,1.0\n", encoding="utf-8")
    (prod / "SubB" / "T.csv").write_text("!2,Age,Rate\n*,30,2.0\n", encoding="utf-8")

    # Minimal Control.xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Config"
    ws.append(["key", "value"])
    ws.append(["working_root", str(root)])
    ws.append(["mode", "generate_changelog"])
    ws.append(["production_tables_path", "Production_Tables"])
    ws.append(["output_path", "Output"])
    ws.append(["run_id", "SUBTEST"])
    ws_cr = wb.create_sheet("ChangeRequests")
    ws_cr.append(
        ["change_request_id", "order", "include", "approved", "description", "notes"]
    )
    ws_cr.append([cr, 1, "Y", "Y", "subfolder test", ""])
    wb.create_sheet("ColumnRenames").append(
        ["change_request_id", "table_name", "old_column_name", "new_column_name"]
    )
    wb.create_sheet("KeyCountApprovals").append(
        ["change_request_id", "table_name", "approved"]
    )
    control_path = root / "Control.xlsx"
    wb.save(control_path)

    clog = generate_change_log(control_path)
    detail_csv = clog.with_name(f"{clog.stem}_Detail.csv")
    assert detail_csv.is_file()
    text = detail_csv.read_text(encoding="utf-8")
    assert "SubA/T" in text
    assert "value_update" in text
    # SubB unchanged → no detail rows, but Summary still lists both tables_touched
    wb_sum = load_workbook(clog, data_only=True)
    summary = list(wb_sum["Summary"].iter_rows(values_only=True))
    wb_sum.close()
    header = summary[0]
    touched_idx = list(header).index("tables_touched")
    touched = str(summary[1][touched_idx])
    assert "SubA/T" in touched
    assert "SubB/T" in touched
    lines = [ln for ln in text.splitlines() if "SubB/T" in ln and "value_update" in ln]
    assert lines == []

    # Review sheets get unique sanitized names
    wb2 = load_workbook(clog)
    assert "ChangeLog_Detail" not in wb2.sheetnames
    assert any("SubA" in n for n in wb2.sheetnames)
    wb2.close()


def _cr(
    change_request_id: str,
    change_type: str,
    *,
    key_tuple: str = "",
    column_name: str = "",
    old_value: str = "",
    new_value: str = "",
    table_name: str = "MORT_TABLE",
    notes: str = "",
) -> ChangeRow:
    return ChangeRow(
        change_request_id=change_request_id,
        table_name=table_name,
        change_type=change_type,
        n_keys_before=2,
        n_keys_after=2,
        key_tuple=key_tuple,
        column_name=column_name,
        old_value=old_value,
        new_value=new_value,
        notes=notes,
    )


def test_missing_row_column_fill_detected():
    """CR_A adds a row; CR_B adds a column — intersection has no covering cell."""
    rows = [
        _cr("CR_A", "row_add", key_tuple="99", column_name="Age", new_value="99"),
        _cr("CR_A", "row_add", key_tuple="99", column_name="Rate", new_value="0.5"),
        _cr("CR_B", "column_add", column_name="NewCol"),
        _cr(
            "CR_B",
            "value_update",
            key_tuple="20",
            column_name="NewCol",
            new_value="x",
            notes="Value on newly added column",
        ),
    ]
    conflicts = detect_conflicts(rows)
    miss = [c for c in conflicts if c["conflict_type"] == "missing_row_column_fill"]
    assert len(miss) == 1
    assert miss[0]["table_name"] == "MORT_TABLE"
    assert miss[0]["key_tuple"] == "99"
    assert miss[0]["column_name"] == "NewCol"
    assert miss[0]["change_request_ids"] == ["CR_A", "CR_B"]
    assert miss[0]["resolved"] == "N"
    assert "without a value" in miss[0]["notes"]


def test_missing_row_column_fill_covered_by_value_update():
    """Explicit covering value_update (even empty) clears the gap."""
    rows = [
        _cr("CR_A", "row_add", key_tuple="99", column_name="Age", new_value="99"),
        _cr("CR_A", "row_add", key_tuple="99", column_name="Rate", new_value="0.5"),
        _cr("CR_B", "column_add", column_name="NewCol"),
        _cr("CR_B", "value_update", key_tuple="99", column_name="NewCol", new_value=""),
    ]
    conflicts = detect_conflicts(rows)
    assert not any(c["conflict_type"] == "missing_row_column_fill" for c in conflicts)


def test_missing_row_column_fill_same_cr_no_false_positive():
    """Same CR row_add cells already include the new column — not flagged."""
    rows = [
        _cr("CR_A", "column_add", column_name="NewCol"),
        _cr("CR_A", "row_add", key_tuple="99", column_name="Age", new_value="99"),
        _cr("CR_A", "row_add", key_tuple="99", column_name="Rate", new_value="0.5"),
        _cr("CR_A", "row_add", key_tuple="99", column_name="NewCol", new_value="y"),
        _cr(
            "CR_A",
            "value_update",
            key_tuple="20",
            column_name="NewCol",
            new_value="x",
            notes="Value on newly added column",
        ),
    ]
    conflicts = detect_conflicts(rows)
    assert not any(c["conflict_type"] == "missing_row_column_fill" for c in conflicts)
